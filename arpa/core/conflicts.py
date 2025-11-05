import pandas as pd
import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers

def extract_specific_columns(input_file, output_file, custom_headers=None, year=None):
    try:
        # Setup output directory
        os.makedirs(os.path.dirname(output_file), exist_ok=True)

        # Read raw data (no automatic parsing)
        df = pd.read_excel(input_file, header=None)

        # Check initial number of rows in the raw data (starting from row 4, which is index 3)
        initial_raw_rows = df.shape[0] - 3
        print(f"Initial raw data rows (after header rows): {initial_raw_rows}")

        # Column selection (first 11 + specified extras)
        base_cols = list(range(11))  # Columns 0-10 (A-K)
        # Add the new detail columns
        extra_cols = [11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29]
        selected_cols = [col for col in base_cols + extra_cols if col < df.shape[1]]

        # This operation itself selects all rows from index 3 onwards for the selected columns
        result = df.iloc[3:, selected_cols].copy()
        result.columns = df.iloc[2, selected_cols].values

        print(f"Rows after initial column selection and header application: {result.shape[0]}")

        # Apply custom headers if provided
        if custom_headers is not None:
            if len(custom_headers) != len(result.columns):
                # Revisit this check if 'Año' is added directly into custom_headers list
                raise ValueError(f"Custom headers count ({len(custom_headers)}) doesn't match column count ({len(result.columns)})")
            result.columns = custom_headers
            print(f"Columns after applying custom headers: {result.columns.tolist()}")

        # Add the 'Año' column to the result DataFrame
        if year is not None:
            result['Año'] = year
        else:
            try:
                filename_without_ext = os.path.basename(input_file).split('.')[0]
                year_from_filename = int("".join(filter(str.isdigit, filename_without_ext)))
                result['Año'] = year_from_filename
                print(f"Deduced year from filename: {year_from_filename}")
            except ValueError:
                print("Warning: Could not extract year from filename and no year was provided. 'Año' column will be empty.")
                result['Año'] = pd.NA # Or a default value if preferred

        # Ensure 'Nombre' concatenation handles all rows
        primer_nombre_col_idx = 3 
        primer_apellido_col_idx = 4 
        segundo_apellido_col_idx = 5 

        if primer_nombre_col_idx < df.shape[1] and primer_apellido_col_idx < df.shape[1] and segundo_apellido_col_idx < df.shape[1]:
            temp_df_for_name = df.iloc[3:, [2, 3, 4, 5]].copy()
            temp_df_for_name = temp_df_for_name.fillna('')
            result_nombre_series = temp_df_for_name.iloc[:, 0].astype(str) + " " + \
                                   temp_df_for_name.iloc[:, 1].astype(str) + " " + \
                                   temp_df_for_name.iloc[:, 2].astype(str) + " " + \
                                   temp_df_for_name.iloc[:, 3].astype(str)
            # Ensure the 'Nombre' column is assigned based on the index of 'result'
            result["Nombre"] = result_nombre_series.values # Assign values directly to avoid index alignment issues if result has non-contiguous index
            print(f"Rows after 'Nombre' concatenation: {result.shape[0]}")
        else:
            print("Warning: Not all name columns (C, D, E, F) found in the input DataFrame for name concatenation.")


        # Process "Nombre" column AFTER merging
        if "Nombre" in result.columns:
            result["Nombre"] = result["Nombre"].fillna("")
            result["Nombre"] = result["Nombre"].replace(r'(?i)\bNan\b', '', regex=True)
            result["Nombre"] = result["Nombre"].str.replace(r'\s+', ' ', regex=True).str.strip()
            result["Nombre"] = result["Nombre"].str.title()
            print(f"Rows after 'Nombre' cleanup: {result.shape[0]}")

        # Replace empty strings with pd.NA (NaN)
        result.replace('', pd.NA, inplace=True)
        print(f"Rows after replacing empty strings with NA: {result.shape[0]}")


        # Special handling for Column J (input index 9), which becomes 'Fecha de Inicio' in custom headers
        if "Fecha de Inicio" in result.columns:
            date_col = "Fecha de Inicio"

            result[date_col] = pd.to_datetime(
                result[date_col],
                dayfirst=True,
                errors='coerce'
            )
            print(f"Rows after date conversion: {result.shape[0]}")

            # Save with Excel formatting
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                result.to_excel(writer, index=False)

                worksheet = writer.sheets['Sheet1']
                date_col_letter = get_column_letter(result.columns.get_loc(date_col) + 1)

                for cell in worksheet[date_col_letter]:
                    if cell.row == 1:
                        continue
                    cell.number_format = 'DD/MM/YYYY'

                for idx, col in enumerate(result.columns):
                    col_letter = get_column_letter(idx+1)
                    worksheet.column_dimensions[col_letter].width = max(
                        len(str(col))+2,
                        (result[col].astype(str).str.len().max() or 0) + 2
                    )
            print(f"Successfully saved '{output_file}' with {result.shape[0]} rows.")
        else:
            print("Warning: 'Fecha de Inicio' column not found in processed data. Saving without date formatting.")
            result.to_excel(output_file, index=False)
            print(f"Successfully saved '{output_file}' with {result.shape[0]} rows (no date formatting).")

        return result

    except Exception as e:
        print(f"Error in extract_specific_columns: {str(e)}")
        return pd.DataFrame()

def generate_justrue_file(input_df, output_file):
    try:
        os.makedirs(os.path.dirname(output_file), exist_ok=True)

        justrue_data = pd.DataFrame()

        if 'Cedula' in input_df.columns:
            justrue_data['Cedula'] = input_df['Cedula']
        else:
            print("Error: 'Cedula' column not found in the input DataFrame for jusTrue file generation.")
            return

        if 'Año' in input_df.columns:
            justrue_data['Año'] = input_df['Año']

        q_columns = [f"Q{i}" for i in range(1, 8)] + [f"Q{i}" for i in range(10, 12)]

        for q_col in q_columns:
            if q_col in input_df.columns:
                # Assign 'true' where the condition is met, keeping original index alignment
                # Use .loc for setting values to avoid SettingWithCopyWarning
                justrue_data[q_col] = input_df[q_col].apply(lambda x: 'true' if str(x).lower() == 'true' else pd.NA)
            else:
                print(f"Warning: Column '{q_col}' not found in the input DataFrame for jusTrue file generation.")

        cols_to_check = [f"Q{i}" for i in range(1, 8)]

        initial_justrue_rows = justrue_data.shape[0]
        justrue_data.dropna(subset=cols_to_check, how='all', inplace=True)
        print(f"Rows in jusTrue data before dropping NAs: {initial_justrue_rows}")
        print(f"Rows in jusTrue data after dropping NAs (only rows with at least one Q1-Q7 'true'): {justrue_data.shape[0]}")


        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            justrue_data.to_excel(writer, index=False)

            worksheet = writer.sheets['Sheet1']
            for idx, col in enumerate(justrue_data.columns):
                col_letter = get_column_letter(idx + 1)
                worksheet.column_dimensions[col_letter].width = max(
                    len(str(col)) + 2,
                    (justrue_data[col].astype(str).str.len().max() or 0) + 2
                )

        print(f"Successfully created '{output_file}' with filtered data.")

    except Exception as e:
        print(f"Error creating jusTrue file: {str(e)}")

# 'Año' will be added dynamically, so it's not in this list.
custom_headers = [
    "ID", "Cedula", "Nombre", "1er Nombre", "1er Apellido",
    "2do Apellido", "Compañía", "Cargo", "Email", "Fecha de Inicio",
    "Q1", "Q1 Detalle", "Q2", "Q2 Detalle", "Q3", "Q3 Detalle",
    "Q4", "Q4 Detalle", "Q5", "Q5 Detalle", "Q6", "Q6 Detalle",
    "Q7", "Q7 Detalle", "Q8", "Q9", "Q10", "Q10 Detalle", "Q11", "Q11 Detalle"
]

# Assuming current year is 2024 for the conflictos.xlsx file.
current_year = 2024

processed_df = extract_specific_columns(
    input_file="core/src/conflictos.xlsx",
    output_file="core/src/conflicts.xlsx",
    custom_headers=custom_headers,
    year=current_year # Pass the year explicitly
)

# Then, if the processing was successful, generate the jusTrue.xlsx file
if not processed_df.empty:
    generate_justrue_file(
        input_df=processed_df,
        output_file="core/src/jusTrue.xlsx"
    )
